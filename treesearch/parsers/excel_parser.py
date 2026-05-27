# -*- coding: utf-8 -*-
"""
@author:XuMing(xuming624@qq.com)
@description: Excel/Spreadsheet parser for TreeSearch.

Requires optional dependency: ``pip install openpyxl``
Extracts sheets, headers, and row data from Excel files and builds tree structure.
"""
import logging
import os
from typing import Optional

logger = logging.getLogger(__name__)

# Extensions supported by openpyxl
EXCEL_EXTENSIONS = frozenset({".xlsx", ".xlsm", ".xltx", ".xltm"})


def _extract_excel_data(
    excel_path: str,
    *,
    max_rows_per_sheet: int = 10000,
    max_consecutive_empty_rows: int = 100,
) -> list[dict]:
    """Extract sheet data from an Excel file.

    Returns a flat node list with:
    - Level 1: Sheet name
    - Level 2: Header row (columns)
    - Level 3: Data rows

    Args:
        excel_path: Path to the Excel file.
        max_rows_per_sheet: Maximum number of rows to process per sheet.
        max_consecutive_empty_rows: Stop parsing a sheet after this many
            consecutive empty rows.
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise ImportError(
            "Excel support requires 'openpyxl'. Install with: pip install openpyxl"
        )

    wb = load_workbook(excel_path, read_only=True, data_only=True)
    nodes = []
    row_counter = 1

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows_iter = ws.iter_rows(values_only=True)

        # Read header row first
        try:
            header_row = next(rows_iter)
        except StopIteration:
            nodes.append({
                "title": sheet_name,
                "level": 1,
                "text": "(empty sheet)",
                "line_num": row_counter,
                "line_start": row_counter,
                "line_end": row_counter,
            })
            row_counter += 1
            continue

        headers = [str(cell) if cell is not None else "" for cell in header_row]
        header_text = f"Columns: {', '.join(h for h in headers if h)}"

        # Sheet node (level 1)
        sheet_start = row_counter
        sheet_text_parts = [header_text]

        # Collect data rows with early termination
        data_rows_text = []
        consecutive_empty = 0
        rows_processed = 0
        truncated = False

        for row in rows_iter:
            rows_processed += 1
            if rows_processed > max_rows_per_sheet:
                truncated = True
                break

            cells = [str(cell) if cell is not None else "" for cell in row]
            is_empty = not any(c.strip() for c in cells)

            if is_empty:
                consecutive_empty += 1
                if consecutive_empty >= max_consecutive_empty_rows:
                    truncated = True
                    break
                continue

            # Reset counter on non-empty row
            consecutive_empty = 0

            row_text = "; ".join(
                f"{h}: {v}" for h, v in zip(headers, cells) if v.strip()
            )
            if row_text:
                data_rows_text.append(row_text)

        # Combine sheet content
        if data_rows_text:
            # Show up to 200 rows in text to avoid excessive node size
            displayed_rows = data_rows_text[:200]
            sheet_text_parts.extend(displayed_rows)
            if len(data_rows_text) > 200:
                sheet_text_parts.append(f"... ({len(data_rows_text) - 200} more rows)")

        if truncated:
            sheet_text_parts.append(
                f"(parsing stopped: reached limit of {max_rows_per_sheet} rows "
                f"or {max_consecutive_empty_rows} consecutive empty rows)"
            )

        sheet_text = "\n".join(sheet_text_parts)
        row_count = len(data_rows_text)

        nodes.append({
            "title": f"{sheet_name} ({row_count} rows)",
            "level": 1,
            "text": sheet_text,
            "line_num": sheet_start,
            "line_start": sheet_start,
            "line_end": sheet_start + row_count,
        })
        row_counter += row_count + 1

    wb.close()
    return nodes


async def excel_to_tree(
    excel_path: str,
    *,
    model: Optional[str] = None,
    if_add_node_summary: bool = True,
    summary_chars_threshold: int = 600,
    if_add_doc_description: bool = False,
    if_add_node_text: bool = False,
    if_add_node_id: bool = True,
    **kwargs,
) -> dict:
    """Build a tree index from an Excel file.

    Each worksheet becomes a level-1 node. Headers and row data
    are extracted as structured text content.

    Returns:
        {'doc_name': str, 'structure': list, 'source_path': str}
    """
    from ..config import get_config

    doc_name = os.path.splitext(os.path.basename(excel_path))[0]
    logger.debug("Parsing Excel: %s", excel_path)

    cfg = get_config()
    nodes = _extract_excel_data(
        excel_path,
        max_rows_per_sheet=cfg.xlsx_max_rows_per_sheet,
        max_consecutive_empty_rows=cfg.xlsx_max_consecutive_empty_rows,
    )

    if not nodes:
        # Empty workbook, create a single root node
        nodes = [{"title": doc_name, "level": 1, "text": "(empty workbook)",
                  "line_num": 1, "line_start": 1, "line_end": 1}]

    from ..indexer import _build_tree, _finalize_tree

    tree = _build_tree(nodes)

    return _finalize_tree(
        tree, doc_name,
        source_path=os.path.abspath(excel_path),
        source_type="excel",
        if_add_node_id=if_add_node_id,
        if_add_node_summary=if_add_node_summary,
        summary_chars_threshold=summary_chars_threshold,
        if_add_node_text=if_add_node_text,
        if_add_doc_description=if_add_doc_description,
    )
